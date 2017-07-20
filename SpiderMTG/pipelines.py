# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import scrapy
import logging
from openpyxl import Workbook
from SpiderMTG.items import Auction


class ExcelWriter(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.width_adjust = 1.2

    def open_spider(self, spider):
        self.ws.cell(row=1, column=1, value='Card PT')
        self.ws.cell(row=1, column=2, value='Card ENG')
        self.ws.cell(row=1, column=3, value='Price')
        self.ws.cell(row=1, column=4, value='Language')
        self.ws.cell(row=1, column=5, value='Expansion')
        self.ws.cell(row=1, column=6, value='Quantity')
        self.ws.cell(row=1, column=7, value='Time Left')
        self.ws.cell(row=1, column=8, value='Condition')
        self.ws.cell(row=1, column=9, value='Link')
        self.ws.cell(row=1, column=10, value='Extra')
        self.ws.cell(row=1, column=11, value='Bids')
        self.ws.cell(row=1, column=12, value='Title')

    def close_spider(self, spider):
        self.update_width()
        try:
            self.wb.save('auctions.xlsx')
        except Exception as e:
            logging.error("Error saving workbook")
            logging.error(e.with_traceback())

    def update_width(self):
        for column_cells in self.ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells) * self.width_adjust
            self.ws.column_dimensions[column_cells[0].column].width = length

    def check_width(self, i, value):
        width = len(value) + self.width_offset
        if width > self.max_cell_width[i]:
            self.max_cell_width[i] = width

    def write_to_workbook(self, item):
        row = self.ws[self.ws.max_row + 1]

        card = item['card']
        if len(card) > 1:
            card_pt = card[0]
            card_eng = card[1]
        else:
            card_pt = ""
            card_eng = card[0]

        row[0].value = card_pt
        row[1].value = card_eng
        row[2].value = str(item['price'])
        row[3].value = str(item['language'])

        exp = item['expansion']
        if len(exp) > 1:
            exp_str = "{pt} / {eng}".format(pt=exp[0], eng=exp[1])
        else:
            exp_str = exp[0]

        row[4].value = exp_str
        row[5].value = str(item['quantity'])
        row[6].value = str(item['time_left'])
        row[7].value = str(item['condition'])
        row[8].value = str(item['href'])
        try:
            row[9].value = str(item['extra'])
        except KeyError as e:
            # There is no extra information
            pass
        row[10].value = str(item['bids'])

        row[11].value = str(item['title'])

    def process_item(self, item, spider):
        self.write_to_workbook(item)

        return item
